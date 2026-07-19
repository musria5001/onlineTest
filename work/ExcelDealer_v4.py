import openpyxl,re,xlrd,xlwt
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment
from openpyxl.comments import Comment
from openpyxl.utils.exceptions import *
from openpyxl.utils import get_column_letter

class ExcelDealer:
    def __init__(self, path_s, path_t, path_r):
        """初始化"""
        # 异常记录
        self.log = set()
        # 路径
        self.path_s = path_s
        self.path_t = path_t
        self.path_r = path_r #+ r"\result.xlsx"
        # 数据存放字典
        self.source_data = {}
        self.target_data = {}
        self.unsure_data = {}
        # 样式
        self.lv4 = NamedStyle(name="lv4")
        self.lv4.fill = PatternFill("solid", fgColor="99FFAA") # 绿
        self.lv4_ = NamedStyle(name="lv4_")
        self.lv4_.fill = PatternFill("solid", fgColor="59BF80") # 暗绿
        self.lv3 = NamedStyle(name="lv3")
        self.lv3.fill = PatternFill("solid", fgColor="66FFEE") # 青
        self.lv3_ = NamedStyle(name="lv3_")
        self.lv3_.fill = PatternFill("solid", fgColor="4DBFB3") # 暗青
        self.lv2 = NamedStyle(name="lv2")
        self.lv2.fill = PatternFill("solid", fgColor="FFE396") # 橙
        self.lv2_ = NamedStyle(name="lv2_")
        self.lv2_.fill = PatternFill("solid", fgColor="FFCC99") # 暗橙
        self.lv1 = NamedStyle(name="lv1")
        self.lv1.fill = PatternFill("solid", fgColor="FF0000") # 红
        self.lv0 = NamedStyle(name="lv0")
        self.lv0.fill = PatternFill("solid", fgColor="FFFF33") # 黄
        # 输出结果文件
        self.output_title = ["班级名称", "学号", "姓名", "平时成绩", "期中成绩", "期末成绩", "实验成绩", "备注", "重修标记", "课程性质", "确信等级"]
        self.output_key = ["classId", "Id", "realName", "score", "again", "courseType"]
        self.result = openpyxl.Workbook()
        self.rs = self.result.active
        # 目标名单版本判断记录，默认不是xls
        self.is_xls_s = False
        self.is_xls_t = False

        # 数据源文件预打开
        self.pat_s = ['学号', '学生昵称', '真实姓名', '成绩/']
        self.pat_s_res = [None, None, None, None]
        try:
            self.source = openpyxl.load_workbook(path_s)
            self.ss = self.source.active
            self.row_start_s = 2
            self.row_end_s = self.ss.max_row + 1
            for i in range(1, self.ss.max_column + 1):
                for j in range(4):
                    if re.match(self.pat_s[j], self.ss[get_column_letter(i) + str(1)].value):
                        self.pat_s_res[j] = i
        except InvalidFileException:
            try:
                self.source = xlrd.open_workbook(path_s)
                self.ss = self.source.sheet_by_index(0)
                self.row_start_s = 1
                self.row_end_s = self.ss.nrows
                self.is_xls_s = True
                for i in range(0, self.ss.ncols):
                    for j in range(4):
                        if re.match(self.pat_s[j], self.ss.cell_value(0, i)):
                            self.pat_s_res[j] = i
            except FileNotFoundError:
                self.log.add("nfs")

        # 源文件信息识别标记
        self.id_col_s = self.pat_s_res[0]
        self.nickname_col = self.pat_s_res[1]
        self.real_name_col = self.pat_s_res[2]
        self.score_col = self.pat_s_res[3]

        # 目标名单文件预打开
        self.pat_t = ['学号', '班级', '姓名', '重修标记', '课程性质']
        self.pat_t_res = [None, None, None, None, None]
        try:
            self.target = openpyxl.load_workbook(path_t)
            self.ts = self.target.active
            self.row_start_t = 1
            self.row_end_t = self.ts.max_row + 1
            for i in range(self.row_start_t, self.row_end_t):
                rec = [None, None, None, None, None]
                for col in range(1, self.ts.max_column + 1):
                    for j in range(5):
                        if re.search(self.pat_t[j], self.ts[get_column_letter(col) + str(i)].value):
                            rec[j] = [i, col + 1]
                    if rec[0] is not None and rec[1] is not None and rec[2] is not None:
                        self.pat_t_res = rec
        except InvalidFileException:
            try:
                self.target = xlrd.open_workbook(path_t)
                self.ts = self.target.sheet_by_index(0)
                self.row_start_t = 0
                self.row_end_t = self.ts.nrows
                self.is_xls_t = True
                for i in range(self.row_start_t, self.row_end_t):
                    rec = [None, None, None, None, None]
                    for col in range(0, self.ts.ncols):
                        for j in range(5):
                            if re.search(self.pat_t[j], str(self.ts.cell_value(i, col))):
                                rec[j] = [i, col]
                        if rec[0] is not None and rec[1] is not None and rec[2] is not None:
                            self.pat_t_res = rec
            except FileNotFoundError:
                self.log.add("nft")

        # 目标文件信息识别标记
        self.row_start_t += self.pat_t_res[0][0] + 1
        self.id_col_t = self.pat_t_res[0][1]
        self.class_id_col = self.pat_t_res[1][1]
        self.name_col = self.pat_t_res[2][1]
        if self.pat_t_res[3]:
            self.again_col = self.pat_t_res[3][1]
        else:
            self.again_col = None
        if self.pat_t_res[4]:
            self.type_col = self.pat_t_res[4][1]
        else:
            self.type_col = None

    def get_data_s(self):
        """慕课数据文件信息提取"""
        for row in range(self.row_start_s, self.row_end_s):
            row_str = str(row)
            if self.is_xls_s:
                """xls"""
                nickname = str(self.ss.cell_value(row, self.nickname_col))
                real_name = str(self.ss.cell_value(row, self.real_name_col))
                id_s = str(self.ss.cell_value(row, self.id_col_s))
                score = int(self.ss.cell_value(row, self.score_col) + 0.5)
            else:
                nickname = str(self.ss[get_column_letter(self.nickname_col) + row_str].value)
                real_name = str(self.ss[get_column_letter(self.real_name_col) + row_str].value)
                id_s = str(self.ss[get_column_letter(self.id_col_s) + row_str].value)
                score = int(self.ss[get_column_letter(self.score_col) + row_str].value + 0.5)
            # 数据记录
            self.source_data[row_str] = {'nickname': nickname, 'realName': real_name, 'Id': id_s, 'score': score}

    def get_data_t(self):
        """目标名单文件信息提取"""
        for row in range(self.row_start_t, self.row_end_t):
            if self.is_xls_t:
                row_str = str(row + 1)
                class_id = str(self.ts.cell_value(row, self.class_id_col))
                id_t = str(self.ts.cell_value(row, self.id_col_t))
                real_name = str(self.ts.cell_value(row, self.name_col))
                if self.again_col is not None:
                    again = self.ts.cell_value(row, self.again_col)
                else:
                    again = ""
                if self.type_col is not None:
                    course_type = self.ts.cell_value(row, self.type_col)
                else:
                    course_type = ""
            else:
                row_str = str(row)
                class_id = str(self.ts[get_column_letter(self.class_id_col) + row_str].value)
                id_t = str(self.ts[get_column_letter(self.id_col_t) + row_str].value)
                real_name = str(self.ts[get_column_letter(self.name_col) + row_str].value)
                if self.again_col is not None:
                    again = self.ts[get_column_letter(self.again_col) + row_str].value
                else:
                    again = ""
                if self.type_col is not None:
                    course_type = self.ts[get_column_letter(self.type_col) + row_str].value
                else:
                    course_type = ""
            # 跳过空数据
            if class_id == 'None' and id_t == 'None' and real_name == 'None':
                continue
            # 数据记录
            self.target_data[row_str] = {'classId': class_id, 'realName': real_name, 'score': None,
                                         'Id': id_t, 'again': again, 'courseType': course_type, "level": 0}

    def find(self):
        """信息对比查询"""
        for stu in self.target_data:
            self.unsure_data[stu] = []
            for row in self.source_data:
                level = 0
                if self.target_data[stu]['Id'] == self.source_data[row]['Id']:
                    level += 1000
                if re.search(self.target_data[stu]['Id'], self.source_data[row]['nickname']):
                    level += 100
                if self.target_data[stu]['realName'] == self.source_data[row]['realName']:
                    level += 10
                if re.search(self.target_data[stu]['realName'], self.source_data[row]['nickname']):
                    level += 1
                if level > 0:
                    temp = self.source_data[row]
                    temp['level'] = level
                    self.unsure_data[stu].append(temp)
                    if level > self.target_data[stu]['level']:
                        self.target_data[stu]['score'] = self.source_data[row]['score']
                        self.target_data[stu]['level'] = level

    def out_put(self):
        """结果输出"""
        def get_key(key):
            return key['level']
        for col in range(1, len(self.output_title) + 1):
            """打印列标题"""
            self.rs[get_column_letter(col) + "1"] = self.output_title[col - 1]
            self.rs[get_column_letter(col) + "1"].font = Font(bold=True)
        self.rs['M1'].alignment = Alignment(wrap_text=True, vertical="center")
        self.rs['M1'].font = Font(name="宋体", bold=True, color='FF3333')
        self.rs['M1'] = '确信等级：\n高：学号姓名栏符合，昵称中也能匹配到学号姓名 ' \
                        '\n\n较高：学号符合，昵称中能匹配到学号，但姓名不对或昵称中未匹配到相应姓名'\
                        '\n\n中：学号和昵称中的学号匹配只有一项符合' \
                        '\n\n较低：学号不符且昵称中未查询到学号，考虑到重名的可能建议人工审查 '\
                        '\n\n符合程度：\n千位代表学号是否匹配，百位代表昵称中能否找到学号，十位代表姓名是否匹配，个位代表昵称中能否找到姓名'\
                        '\n\n另：学生成绩匹配到的所有可能数据会以批注形式在确信等级上给出，若一名学生成绩包含两条及以上高可能数据（确信等级“中”以上）的会用深色标记'
        self.rs.merge_cells('M1:Q21')
        for row in range(self.row_start_t + 1, len(self.target_data) + 1 + self.row_start_t):
            row_str1 = str(row)
            row_str = str(row - self.row_start_t + 1)
            """打印每列数据"""
            for i in range(1, len(self.output_key) + 1):
                if i > 4:
                    self.rs[get_column_letter(i + 4) + row_str] = self.target_data[row_str1][self.output_key[i - 1]]
                else:
                    self.rs[get_column_letter(i) + row_str] = self.target_data[row_str1][self.output_key[i - 1]]
            comment = ""
            check = 0
            score_h = self.target_data[row_str1]['score']
            self.unsure_data[row_str1].sort(reverse=True, key=get_key)
            for item in self.unsure_data[row_str1]:
                comment += "昵称：" + item['nickname'] + " 真实姓名：" + item['realName'] + " 学号：" + item['Id'] + " 成绩：" + str(item['score']) + " 符合程度：" + str(item['level']) +" \n"
                if item['level'] >= 100:
                    if item['score'] > score_h:
                        score_h = item['score']
                        self.rs['D' + row_str] = score_h
                    check += 1
            level = self.target_data[row_str1]['level']
            if level == 1111:
                self.rs["k" + row_str] = "高"
                if check >= 2:
                    self.rs["k" + row_str].style = self.lv4_
                else:
                    self.rs["k" + row_str].style = self.lv4
            elif level >= 1100:
                self.rs["k" + row_str] = "较高"
                if check >= 2:
                    self.rs["k" + row_str].style = self.lv3_
                else:
                    self.rs["k" + row_str].style = self.lv3
            elif level >= 100:
                if check >= 2:
                    self.rs["k" + row_str] = "中,建议核查"
                    self.rs["k" + row_str].style = self.lv2_
                else:
                    self.rs["k" + row_str] = "中"
                    self.rs["k" + row_str].style = self.lv2
            elif level < 100:
                if level == 0:
                    self.rs["k" + row_str] = "未查到"
                    self.rs["k" + row_str].style = self.lv0
                else:
                    self.rs["k" + row_str] = "较低，建议人工核查！"
                    self.rs["k" + row_str].style = self.lv1
            self.rs["k" + row_str].comment = Comment(comment, '', height=150, width=800)
        self.result.save(self.path_r)

    def run(self):
        if self.log:
            if "nfs" in self.log and "nft" in self.log:
                return "nfa"
            if "nfs" in self.log:
                return "nfs"
            if "nft" in self.log:
                return "nft"
        else:
            self.get_data_s()
            self.get_data_t()
            self.find()
            self.out_put()
            return "suc"

